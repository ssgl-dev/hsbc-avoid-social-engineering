from openpyxl import load_workbook
import argparse
import sys
from typing import Optional
import unicodedata

class ExcelHelper:
    """
    Helper to open workbook and provide find_index_for_translation.
    """
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.workbook = load_workbook(excel_path, data_only=True)
        print(f"Workbook '{excel_path}' loaded.")
        self.sheet_name_dict = { 
            "網絡安全及防詐騙資訊中心 - 香港滙豐":"Sheet1",
            "Cyber security and fraud hub - HSBC HK":"Sheet1",
            "防範惡意軟件 | 網絡安全及防詐騙資訊中心 - 香港滙豐":"Sheet2",
            "Malware Safety Measures | Cyber Security and Fraud - HSBC HK":"Sheet2",
            "如何變得更Cyber Smart？ | 網絡安全及防詐騙資訊 - 香港滙豐":"Sheet3",
            "How to Be Cyber Smart? | Cyber Security and Fraud - HSBC HK":"Sheet3",
            "流動裝置設定您要知 | 網絡安全及防詐騙資訊 - 香港滙豐":"Sheet4",
            "What you need to know about mobile device settings | Cyber security and fraud - HSBC HK":"Sheet4",
            "如何安全使用平板電腦 | 網絡安全及防詐騙資訊 - 香港滙豐":"Sheet5",
            "How to use a laptop/tablet securely | Cyber security and fraud - HSBC HK":"Sheet5",
            "開放銀行服務 | 開放銀行API及如何運作 - 香港滙豐":"Sheet6",
            "Open Banking | Open Banking API & How It Works - HSBC HK":"Sheet6",
            "甚麼是「網絡釣魚」及短訊發送人登記制 - 香港滙豐":"Sheet7",   
            "What is \"phishing\" and SMS Sender Registration Scheme - HSBC HK":"Sheet7",
            "如何避免WhatsApp和Instagram帳戶被盜用 - 香港滙豐":"Sheet8",
            "Stay Aware Of WhatsApp And Instagram Scams - HSBC HK":"Sheet8",
            "如何防範「社交工程」陷阱 | 網絡安全及防詐騙資訊 - 香港滙豐":"Sheet9",
            "Avoid social engineering scams | Cyber security and fraud - HSBC HK":"Sheet9",
            "Beware of malicious software | Cyber security and fraud - HSBC HK":"Sheet10",
            "小心惡意軟件 | 網絡安全及防詐騙資訊 - 香港滙豐":"Sheet10",
            "提防信用卡詐騙 | 網絡安全及防詐騙資訊 - 香港滙豐":"Sheet11",
            "Preventing credit card fraud | Cyber security and fraud - HSBC HK":"Sheet11",
            "懷疑受騙？教您處理及舉報 | 網絡安全及防詐騙資訊 - 香港滙豐":"Sheet12",
            "How to report fraud | Cyber security and fraud - HSBC HK":"Sheet12",
            "防詐騙三大法則 | 網絡安全及防詐騙資訊 - 香港滙豐":"Sheet13",
            "Prevent Fraud | Cyber Security And Fraud - HSBC HK":"Sheet13",
            "如何分辨假冒來電？ | 網絡安全及防詐騙資訊 - 香港滙豐":"Sheet14",
            "How to identify bogus calls | Cyber security and fraud - HSBC HK":"Sheet14",
            "求職騙案 | 網絡安全及防詐騙資訊中心 - 香港滙豐":"Sheet15",
            "How To Avoid Job Scams | Cyber Security And Fraud - HSBC HK":"Sheet15",
            "網戀騙局 | 網絡安全及防詐騙資訊 - 香港滙豐":"Sheet16",
            "Romance Scams | Cyber Security And Fraud - HSBC HK":"Sheet16",
            "電騙陷阱 | 網絡安全及防詐騙資訊中心 - 香港滙豐":"Sheet17",
            "Beware Of Phone Scams | Cyber Security And Fraud - HSBC HK":"Sheet17",
            "小心支票詐騙陷阱 | 網絡安全及防詐騙資訊 - 滙豐香港":"Sheet18",
            "Avoiding Cheque Scams | Cyber Security And Fraud - HSBC HK":"Sheet18",
            "投資騙案手法及防騙貼士 | 香港投資詐騙 - 香港滙豐":"Sheet19",
            "How Investment Scams Work And How To Avoid Them - HSBC HK":"Sheet19",
            "自動櫃員機有甚麼陷阱？ | 網絡安全及防詐騙資訊 - 香港滙豐":"Sheet20",
            "How to avoid ATM traps | Cyber security and fraud - HSBC HK":"Sheet20",
            "如何設定高強度密碼？ | 網絡安全及防詐騙資訊 - 香港滙豐":"Sheet21",
            "How to set a strong password | Cyber security and fraud - HSBC HK":"Sheet21",
            "旅遊有清單，保安一樣有 | 網絡安全及防詐騙資訊 - 香港滙豐":"Sheet22",
            "Security checklist before your next travel departure | Cyber security and fraud - HSBC HK":"Sheet22",
            "HSBC Safeguard：同心打擊金融罪案 | 網絡安全及防詐騙資訊 - 香港滙豐":"Sheet23",
            "HSBC Safeguard: Let's fight financial crime | Cyber security and fraud - HSBC HK":"Sheet23"
        }

    def __enter__(self):
        """context entrance"""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """context exit"""
        self.close()
        return False

    def __del__(self):
        """Make sure excel file closes"""
        if hasattr(self, 'workbook') and self.workbook:
            try:
                self.workbook.close()
            except:
                pass  

    def close(self):
        """Close excel"""
        if self.workbook:
            try:
                self.workbook.close()
                print(f"Workbook '{self.excel_path}' closed successfully.")
            except Exception as e:
                print(f"Error closing workbook '{self.excel_path}': {e}")
            finally:
                self.workbook = None

    @staticmethod
    def _is_space_like(ch: str) -> bool:
        if not ch:
            return False
        if ch.isspace():
            return True
        cat = unicodedata.category(ch)
        if cat.startswith('Z'):
            return True
        # also treat some invisible/zero-width and non-breaking spaces as space-like
        # and trim common Chinese end-of-sentence punctuation
        if ch in ('\u200b', '\ufeff', '\u200c', '\u200d', '\u2060', '\u202f', '\u00a0', '\u2009', '\uff0c', '\u3002', '\uff1b'):
            return True
        return False

    @classmethod
    def full_strip(cls, s: Optional[str]) -> Optional[str]:
        if s is None:
            return None
        start = 0
        end = len(s) - 1
        while start <= end and cls._is_space_like(s[start]):
            start += 1
        while end >= start and cls._is_space_like(s[end]):
            end -= 1
        return s[start:end+1]

    @classmethod
    def normalize_text(cls, text: Optional[str]) -> Optional[str]:
        """Normalize text by removing special characters and performing Unicode normalization"""
        if text is None:
            return None
        
        # First strip whitespace
        text = cls.full_strip(text)
        if not text:
            return text
            
        # Remove trademark symbols and other problematic characters
        text = text.replace('™', '').replace('®', '').replace('©', '')
        
        # Remove zero-width characters
        text = text.replace('\u200b', '').replace('\ufeff', '').replace('\u200c', '').replace('\u200d', '')
        
        # Normalize Unicode (NFKC combines compatibility normalization)
        text = unicodedata.normalize('NFKC', text)
        
        # Final strip after normalization
        return cls.full_strip(text)

    def find_index_for_translation(self, sheet_name: str, text: str) -> Optional[str]:
        if not sheet_name.startswith("Sheet"):
            raise ValueError("sheet_name 格式无效：必须以 'Sheet' 开头")
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"未找到工作表: {sheet_name}")

        ws = self.workbook[sheet_name]
        translation_col = None
        index_col = None
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
        for cell in first_row:
            if cell.value and isinstance(cell.value, str):
                v = cell.value.strip().lower()
                if v == "translation":
                    translation_col = cell.column
                elif v == "index":
                    index_col = cell.column

        if translation_col is None:
            translation_col = 2
        if index_col is None:
            index_col = 1

        target = self.normalize_text(text) or ""
        max_row = ws.max_row

        for r in range(2, max_row + 1):
            if (r - 2) % 4 not in (0, 2):
                continue
            trans_cell = ws.cell(row=r, column=translation_col)
            val = trans_cell.value
            if val is None:
                continue
            if not isinstance(val, str):
                val = str(val)
            cell_text = self.normalize_text(val) or ""
            if cell_text == target:
                idx_row = (r - 2) // 4 * 4 + 2
                idx_cell = ws.cell(row=idx_row, column=index_col)
                idx_val = idx_cell.value
                return None if idx_val is None else str(idx_val)
        return None

    def find_sheet_name_by_title(self, title: str) -> Optional[str]:
        # if title not in self.sheet_name_dict:
        #     return "Sheet1"
        return self.sheet_name_dict.get(title, "Sheet1")
        